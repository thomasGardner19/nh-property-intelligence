"""Extract source rows from the FHFA Annual County HPI XLSX workbook."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from .contract import EXPECTED_FIELDS


def extract_records(workbook_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    header_index = None
    headers: list[str] = []
    for index, row in enumerate(rows):
        candidate = [str(value).strip() if value is not None else "" for value in row]
        if "FIPS code" in candidate and "Year" in candidate and "HPI" in candidate:
            header_index = index
            headers = candidate
            break
    if header_index is None:
        raise ValueError("Could not locate FHFA county HPI header row")
    missing = [field for field in EXPECTED_FIELDS if field not in headers]
    if missing:
        raise ValueError(f"FHFA workbook missing expected fields: {missing}")
    records: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if not any(value is not None and str(value).strip() for value in row):
            continue
        record = {header: value for header, value in zip(headers, row, strict=False) if header}
        records.append(record)
    if not records:
        raise ValueError("FHFA workbook contained no data records")
    return records
